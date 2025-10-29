"""
Generador de plantillas Excel para carga de correos de clientes
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


class TemplateGenerator:
    """Genera plantillas Excel para el sistema"""

    @staticmethod
    def crear_plantilla_correos(ruta_salida):
        """
        Crea una plantilla de Excel para cargar correos de clientes

        Args:
            ruta_salida: Ruta donde guardar la plantilla

        Returns:
            Tupla (exito, mensaje)
        """
        try:
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Clientes"

            # Definir encabezados
            encabezados = ['NIT', 'Nombre', 'Email']

            # Estilo para encabezados
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Escribir encabezados
            for col, encabezado in enumerate(encabezados, 1):
                celda = ws.cell(row=1, column=col, value=encabezado)
                celda.font = header_font
                celda.fill = header_fill
                celda.alignment = header_alignment
                celda.border = border

            # Datos de ejemplo
            ejemplos = [
                ['900123456', 'EMPRESA EJEMPLO S.A.S', 'contacto@ejemplo.com'],
                ['80012345', 'COMERCIALIZADORA ABC LTDA', 'ventas@abc.com'],
                ['1234567890', 'Juan Pérez García', 'juan.perez@email.com'],
                ['31234567', 'María López Rodríguez', 'maria.lopez@email.com; contabilidad@empresa.com'],
            ]

            # Escribir ejemplos
            for row_idx, ejemplo in enumerate(ejemplos, 2):
                for col_idx, valor in enumerate(ejemplo, 1):
                    celda = ws.cell(row=row_idx, column=col_idx, value=valor)
                    celda.border = border

                    # Alineación específica por columna
                    if col_idx == 1:  # NIT
                        celda.alignment = Alignment(horizontal='left')
                    elif col_idx == 2:  # Nombre
                        celda.alignment = Alignment(horizontal='left')
                    else:  # Email
                        celda.alignment = Alignment(horizontal='left')

            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 15  # NIT
            ws.column_dimensions['B'].width = 40  # Nombre
            ws.column_dimensions['C'].width = 50  # Email

            # Agregar hoja de instrucciones
            ws_instrucciones = wb.create_sheet(title="📖 Instrucciones")

            instrucciones = [
                ['PLANTILLA DE CORREOS DE CLIENTES', ''],
                ['', ''],
                ['📋 INSTRUCCIONES DE USO:', ''],
                ['', ''],
                ['1. Columna NIT:', 'Ingrese el NIT o cédula del cliente (7-10 dígitos)'],
                ['', '• Puede incluir o no el dígito de verificación'],
                ['', '• Ejemplos válidos: 900123456, 80012345-9, 1234567890'],
                ['', ''],
                ['2. Columna Nombre:', 'Ingrese el nombre o razón social completa del cliente'],
                ['', '• No puede estar vacío'],
                ['', '• Ejemplos: EMPRESA XYZ S.A.S, Juan Pérez García'],
                ['', ''],
                ['3. Columna Email:', 'Ingrese el(los) correo(s) electrónico(s) del cliente'],
                ['', '• Para múltiples correos, sepárelos con punto y coma (;) o coma (,)'],
                ['', '• Ejemplo: ventas@empresa.com; contabilidad@empresa.com'],
                ['', ''],
                ['✅ NOMBRES DE COLUMNAS ACEPTADOS:', ''],
                ['', ''],
                ['NIT:', 'nit, nit_comprador, numero_identificacion, identificacion, num_id'],
                ['Nombre:', 'nombre, nombre_del_comprador, nombre_comprador, nombre_cliente, razon_social, cliente, empresa'],
                ['Email:', 'email, correos, correo, correo_electronico, e-mail, mail'],
                ['', ''],
                ['📝 NOTAS IMPORTANTES:', ''],
                ['', ''],
                ['', '• El sistema detecta automáticamente las columnas (no es sensible a mayúsculas/minúsculas)'],
                ['', '• Puede usar cualquiera de los nombres de columna aceptados'],
                ['', '• Los datos de ejemplo en la hoja "Clientes" son solo ilustrativos'],
                ['', '• Puede eliminar los ejemplos y agregar sus propios datos'],
                ['', '• Asegúrese de no dejar filas vacías entre los datos'],
                ['', ''],
                ['⚠️ VALIDACIONES:', ''],
                ['', ''],
                ['', '• El NIT debe tener entre 7 y 10 dígitos'],
                ['', '• El email debe tener formato válido (ej: usuario@dominio.com)'],
                ['', '• El nombre no puede estar vacío'],
                ['', '• Si un registro tiene errores, será omitido y se mostrará en el reporte'],
            ]

            # Escribir instrucciones
            for row_idx, (titulo, descripcion) in enumerate(instrucciones, 1):
                celda_a = ws_instrucciones.cell(row=row_idx, column=1, value=titulo)
                celda_b = ws_instrucciones.cell(row=row_idx, column=2, value=descripcion)

                # Formato especial para títulos
                if titulo and not descripcion:
                    celda_a.font = Font(name='Arial', size=14, bold=True, color='0070C0')
                elif titulo and '📋' in titulo or '✅' in titulo or '📝' in titulo or '⚠️' in titulo:
                    celda_a.font = Font(name='Arial', size=12, bold=True, color='FF6600')
                elif titulo and ':' in titulo:
                    celda_a.font = Font(name='Arial', size=10, bold=True)

                celda_b.alignment = Alignment(wrap_text=True, vertical='top')

            # Ajustar ancho de columnas de instrucciones
            ws_instrucciones.column_dimensions['A'].width = 25
            ws_instrucciones.column_dimensions['B'].width = 80

            # Hacer que la hoja de Clientes sea la activa
            wb.active = 0

            # Guardar archivo
            wb.save(ruta_salida)

            return True, f"Plantilla creada exitosamente en: {ruta_salida}"

        except Exception as e:
            return False, f"Error al crear plantilla: {str(e)}"

    @staticmethod
    def obtener_nombre_plantilla_default():
        """
        Obtiene el nombre por defecto para la plantilla

        Returns:
            Nombre del archivo
        """
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d")
        return f"Plantilla_Correos_Clientes_{timestamp}.xlsx"
