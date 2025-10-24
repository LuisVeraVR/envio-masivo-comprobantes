"""
Pestaña de Reportes
Genera reportes de envíos por estado
"""

from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QGroupBox,
                             QLabel, QPushButton, QCheckBox, QDateEdit,
                             QMessageBox, QFileDialog)
from PyQt6.QtCore import Qt, QDate
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os


class ReportesTab(QWidget):
    """Pestaña de generación de reportes"""
    
    def __init__(self, config, db, logger):
        super().__init__()
        self.config = config
        self.db = db
        self.logger = logger
        
        self._crear_interfaz()
    
    def _crear_interfaz(self):
        """Crea la interfaz de la pestaña"""
        layout = QVBoxLayout(self)
        
        # Título
        titulo = QLabel("📊 Generación de Reportes")
        titulo.setStyleSheet("font-size: 18px; font-weight: bold; padding: 10px;")
        layout.addWidget(titulo)
        
        # Grupo: Rango de fechas
        grupo_fechas = self._crear_grupo_fechas()
        layout.addWidget(grupo_fechas)
        
        # Grupo: Estados
        grupo_estados = self._crear_grupo_estados()
        layout.addWidget(grupo_estados)
        
        # Botón generar
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        self.btn_generar = QPushButton("📈 Generar Reporte")
        self.btn_generar.setMinimumHeight(40)
        self.btn_generar.setStyleSheet("""
            QPushButton {
                background-color: #28a745; color: white; font-weight: bold;
                border-radius: 5px; padding: 8px 20px; font-size: 14px;
            }
            QPushButton:hover { background-color: #218838; }
        """)
        self.btn_generar.clicked.connect(self._generar_reporte)
        btn_layout.addWidget(self.btn_generar)
        
        layout.addLayout(btn_layout)
        layout.addStretch()
    
    def _crear_grupo_fechas(self):
        """Crea el grupo de selección de fechas"""
        grupo = QGroupBox("📅 Rango de Fechas")
        layout = QHBoxLayout()
        
        layout.addWidget(QLabel("Desde:"))
        self.fecha_desde = QDateEdit()
        self.fecha_desde.setCalendarPopup(True)
        self.fecha_desde.setDate(QDate.currentDate().addMonths(-1))
        layout.addWidget(self.fecha_desde)
        
        layout.addWidget(QLabel("Hasta:"))
        self.fecha_hasta = QDateEdit()
        self.fecha_hasta.setCalendarPopup(True)
        self.fecha_hasta.setDate(QDate.currentDate())
        layout.addWidget(self.fecha_hasta)
        
        layout.addStretch()
        
        grupo.setLayout(layout)
        return grupo
    
    def _crear_grupo_estados(self):
        """Crea el grupo de selección de estados"""
        grupo = QGroupBox("📋 Estados a Incluir")
        layout = QVBoxLayout()
        
        self.chk_enviados = QCheckBox("✅ Enviados")
        self.chk_enviados.setChecked(True)
        layout.addWidget(self.chk_enviados)
        
        self.chk_rebotados = QCheckBox("🔄 Rebotados")
        self.chk_rebotados.setChecked(True)
        layout.addWidget(self.chk_rebotados)
        
        self.chk_bloqueados = QCheckBox("🚫 Bloqueados")
        self.chk_bloqueados.setChecked(True)
        layout.addWidget(self.chk_bloqueados)
        
        self.chk_inexistentes = QCheckBox("❓ Inexistentes")
        self.chk_inexistentes.setChecked(True)
        layout.addWidget(self.chk_inexistentes)
        
        self.chk_errores = QCheckBox("❌ Errores")
        self.chk_errores.setChecked(True)
        layout.addWidget(self.chk_errores)
        
        grupo.setLayout(layout)
        return grupo
    
    def _generar_reporte(self):
        """Genera el reporte en Excel"""
        # Obtener estados seleccionados
        estados_seleccionados = []
        if self.chk_enviados.isChecked():
            estados_seleccionados.append('ENVIADO')
        if self.chk_rebotados.isChecked():
            estados_seleccionados.append('REBOTADO')
        if self.chk_bloqueados.isChecked():
            estados_seleccionados.append('BLOQUEADO')
        if self.chk_inexistentes.isChecked():
            estados_seleccionados.append('INEXISTENTE')
        if self.chk_errores.isChecked():
            estados_seleccionados.append('ERROR')
        
        if not estados_seleccionados:
            QMessageBox.warning(self, "Sin estados", "Debe seleccionar al menos un estado")
            return
        
        # Obtener fechas
        fecha_desde = self.fecha_desde.date().toString("yyyy-MM-dd")
        fecha_hasta = self.fecha_hasta.date().toString("yyyy-MM-dd")
        
        # Crear workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Eliminar hoja por defecto
        
        # Crear una pestaña por estado
        total_registros = 0
        
        for estado in estados_seleccionados:
            envios = self.db.obtener_envios_por_estado(estado, fecha_desde, fecha_hasta)
            
            if not envios:
                continue
            
            ws = wb.create_sheet(title=estado)
            
            # Encabezados
            headers = ['Fecha', 'NIT', 'Cliente', 'Email', 'Archivos', 'Estado', 'Error']
            ws.append(headers)
            
            # Estilo de encabezados
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Datos
            for envio in envios:
                ws.append([
                    envio['fecha_envio'],
                    envio['nit'],
                    envio['nombre_cliente'],
                    envio['email_destino'],
                    envio['cantidad_archivos'],
                    envio['estado'],
                    envio['mensaje_error'] or ''
                ])
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 40
            
            total_registros += len(envios)
        
        if total_registros == 0:
            QMessageBox.information(self, "Sin datos",
                                   "No se encontraron registros en el rango de fechas seleccionado")
            return
        
        # Guardar archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"reporte_comprobantes_{timestamp}.xlsx"
        ruta_reportes = self.config.get("paths.reports", "reportes/")
        ruta_completa = os.path.join(ruta_reportes, nombre_archivo)
        
        try:
            wb.save(ruta_completa)
            
            self.logger.info(f"Reporte generado: {nombre_archivo} ({total_registros} registros)",
                           modulo="ReportesTab")
            
            respuesta = QMessageBox.question(
                self, "Reporte generado",
                f"Reporte generado exitosamente:\n{nombre_archivo}\n\n"
                f"Total de registros: {total_registros}\n\n"
                "¿Desea abrir la carpeta de reportes?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if respuesta == QMessageBox.StandardButton.Yes:
                import subprocess, platform
                if platform.system() == 'Windows':
                    os.startfile(ruta_reportes)
                elif platform.system() == 'Darwin':
                    subprocess.Popen(['open', ruta_reportes])
                else:
                    subprocess.Popen(['xdg-open', ruta_reportes])
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar reporte: {str(e)}")
            self.logger.error(f"Error al generar reporte: {e}", modulo="ReportesTab", exc_info=True)