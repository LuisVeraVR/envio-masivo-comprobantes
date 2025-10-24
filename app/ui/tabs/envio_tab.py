"""
Pestaña de Envío de Comprobantes
Permite cargar Excel, ZIP y enviar correos masivos
"""

from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QGroupBox,
                             QLabel, QPushButton, QFileDialog, QTextEdit,
                             QMessageBox, QProgressBar, QTableWidget, QTableWidgetItem,
                             QLineEdit, QSpinBox, QFormLayout, QDialog)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from app.core.excel_processor import ExcelProcessor
from app.core.zip_handler import ZipHandler
from app.core.email_sender import EmailSender
import os


class EnvioWorker(QThread):
    """Worker thread para envío de correos"""
    progreso = pyqtSignal(int, int, dict, bool, str, str)  # actual, total, cliente, exito, estado, mensaje
    finalizado = pyqtSignal(dict)  # resumen
    
    def __init__(self, clientes_con_archivos, smtp_config, emails_cc, modo_prueba, db, logger):
        super().__init__()
        self.clientes_con_archivos = clientes_con_archivos
        self.smtp_config = smtp_config
        self.emails_cc = emails_cc
        self.modo_prueba = modo_prueba
        self.db = db
        self.logger = logger
    
    def run(self):
        """Ejecuta el envío en segundo plano"""
        sender = EmailSender(self.smtp_config, self.db, self.logger)
        
        def callback_progreso(i, total, cliente, exito, estado, mensaje):
            self.progreso.emit(i, total, cliente, exito, estado, mensaje)
        
        resumen = sender.enviar_lote(
            self.clientes_con_archivos,
            emails_copia=self.emails_cc,
            modo_prueba=self.modo_prueba,
            usuario_operativa=None,
            callback_progreso=callback_progreso
        )
        
        self.finalizado.emit(resumen)


class EnvioTab(QWidget):
    """Pestaña de envío de comprobantes"""
    
    def __init__(self, config, db, logger):
        super().__init__()
        self.config = config
        self.db = db
        self.logger = logger
        
        self.excel_processor = None
        self.zip_handler = None
        self.clientes = []
        self.archivos_por_nit = {}
        
        self._crear_interfaz()
    
    def _crear_interfaz(self):
        """Crea la interfaz de la pestaña"""
        layout = QVBoxLayout(self)
        
        # Título
        titulo = QLabel("📧 Envío de Comprobantes")
        titulo.setStyleSheet("font-size: 18px; font-weight: bold; padding: 10px;")
        layout.addWidget(titulo)
        
        # Grupo: Configuración de Correo (SMTP)
        grupo_smtp = self._crear_grupo_smtp()
        layout.addWidget(grupo_smtp)
        
        # Grupo: Cargar archivos
        grupo_archivos = self._crear_grupo_archivos()
        layout.addWidget(grupo_archivos)
        
        # Grupo: Vista previa
        grupo_preview = self._crear_grupo_preview()
        layout.addWidget(grupo_preview)
        
        # Barra de progreso
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # Log de envíos
        self.txt_log = QTextEdit()
        self.txt_log.setMaximumHeight(150)
        self.txt_log.setReadOnly(True)
        self.txt_log.setPlaceholderText("Aquí se mostrarán los resultados de envío...")
        layout.addWidget(self.txt_log)
        
        # Botón enviar
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        self.btn_enviar = QPushButton("📤 Enviar Comprobantes")
        self.btn_enviar.setMinimumHeight(40)
        self.btn_enviar.setEnabled(False)
        self.btn_enviar.setStyleSheet("""
            QPushButton {
                background-color: #007bff; color: white; font-weight: bold;
                border-radius: 5px; padding: 8px 20px; font-size: 14px;
            }
            QPushButton:hover { background-color: #0056b3; }
            QPushButton:disabled { background-color: #ccc; }
        """)
        self.btn_enviar.clicked.connect(self._enviar_comprobantes)
        btn_layout.addWidget(self.btn_enviar)
        
        layout.addLayout(btn_layout)
    
    def _crear_grupo_smtp(self):
        """Crea el grupo de configuración SMTP"""
        grupo = QGroupBox("🔐 Configuración de Correo (Para enviar desde su cuenta)")
        grupo.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #007bff;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
        """)
        
        from PyQt6.QtWidgets import QFormLayout, QSpinBox, QCheckBox
        
        layout = QFormLayout()
        layout.setSpacing(8)
        
        # Correo
        self.txt_correo = QLineEdit()
        self.txt_correo.setPlaceholderText("su-correo@gmail.com")
        layout.addRow("📧 Su Correo:", self.txt_correo)
        
        # Contraseña
        self.txt_password_envio = QLineEdit()
        self.txt_password_envio.setEchoMode(QLineEdit.EchoMode.Password)
        self.txt_password_envio.setPlaceholderText("Su contraseña")
        
        password_layout = QHBoxLayout()
        password_layout.addWidget(self.txt_password_envio)
        
        self.btn_mostrar_pass = QPushButton("👁")
        self.btn_mostrar_pass.setMaximumWidth(40)
        self.btn_mostrar_pass.setCheckable(True)
        self.btn_mostrar_pass.clicked.connect(self._toggle_password)
        password_layout.addWidget(self.btn_mostrar_pass)
        
        layout.addRow("🔑 Contraseña:", password_layout)
        
        # Servidor SMTP (autodetección)
        servidor_layout = QHBoxLayout()
        self.txt_servidor = QLineEdit()
        self.txt_servidor.setText("smtp.gmail.com")
        self.txt_servidor.setPlaceholderText("smtp.gmail.com")
        servidor_layout.addWidget(self.txt_servidor)
        
        self.spin_puerto = QSpinBox()
        self.spin_puerto.setRange(1, 65535)
        self.spin_puerto.setValue(587)
        self.spin_puerto.setMaximumWidth(80)
        servidor_layout.addWidget(QLabel("Puerto:"))
        servidor_layout.addWidget(self.spin_puerto)
        
        layout.addRow("🖥 Servidor:", servidor_layout)
        
        # Correos en copia (CC)
        self.txt_cc = QLineEdit()
        self.txt_cc.setPlaceholderText("copia1@empresa.com, copia2@empresa.com")
        layout.addRow("📋 Copias (CC):", self.txt_cc)
        
        # Ayuda
        ayuda = QLabel(
            '<small><b>Gmail:</b> Use contraseña de aplicación. '
            '<a href="https://support.google.com/accounts/answer/185833">¿Cómo obtenerla?</a></small>'
        )
        ayuda.setOpenExternalLinks(True)
        ayuda.setStyleSheet("padding: 5px; background-color: #fff3cd; border-radius: 3px;")
        layout.addRow("", ayuda)
        
        # Botón probar
        btn_probar = QPushButton("🔍 Probar Conexión")
        btn_probar.clicked.connect(self._probar_conexion_smtp)
        layout.addRow("", btn_probar)
        
        grupo.setLayout(layout)
        return grupo
    
    def _toggle_password(self):
        """Muestra u oculta la contraseña"""
        if self.btn_mostrar_pass.isChecked():
            self.txt_password_envio.setEchoMode(QLineEdit.EchoMode.Normal)
            self.btn_mostrar_pass.setText("🙈")
        else:
            self.txt_password_envio.setEchoMode(QLineEdit.EchoMode.Password)
            self.btn_mostrar_pass.setText("👁")
    
    def _probar_conexion_smtp(self):
        """Prueba la conexión SMTP"""
        from app.core.email_sender import EmailSender
        from app.utils.validator import Validator
        
        correo = self.txt_correo.text().strip()
        password = self.txt_password_envio.text()
        servidor = self.txt_servidor.text().strip()
        puerto = self.spin_puerto.value()
        
        es_valido, mensaje = Validator.validar_configuracion_smtp(servidor, puerto, correo, password)
        if not es_valido:
            QMessageBox.warning(self, "Configuración inválida", mensaje)
            return
        
        smtp_config = {
            'server': servidor,
            'port': puerto,
            'username': correo,
            'password': password,
            'use_tls': True,
            'from_name': 'Sistema de Comprobantes'
        }
        
        try:
            sender = EmailSender(smtp_config, self.db, self.logger)
            exito, msg = sender.probar_conexion()
            
            if exito:
                QMessageBox.information(self, "Conexión exitosa", 
                    f"✓ {msg}\n\n¡Puede enviar correos!")
            else:
                QMessageBox.critical(self, "Error de conexión", f"✗ {msg}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error: {str(e)}")
    
    def _crear_grupo_archivos(self):
        """Crea el grupo de carga de archivos"""
        grupo = QGroupBox("📁 Cargar Archivos")
        layout = QVBoxLayout()
        
        # Excel
        excel_layout = QHBoxLayout()
        self.label_excel = QLabel("❌ Excel no cargado")
        excel_layout.addWidget(self.label_excel)
        excel_layout.addStretch()
        
        btn_excel = QPushButton("📊 Cargar Excel")
        btn_excel.clicked.connect(self._cargar_excel)
        excel_layout.addWidget(btn_excel)
        
        layout.addLayout(excel_layout)
        
        # ZIP
        zip_layout = QHBoxLayout()
        self.label_zip = QLabel("❌ ZIP no cargado")
        zip_layout.addWidget(self.label_zip)
        zip_layout.addStretch()
        
        btn_zip = QPushButton("📦 Cargar ZIP")
        btn_zip.clicked.connect(self._cargar_zip)
        zip_layout.addWidget(btn_zip)
        
        layout.addLayout(zip_layout)
        
        grupo.setLayout(layout)
        return grupo
    
    def _crear_grupo_preview(self):
        """Crea el grupo de vista previa"""
        grupo = QGroupBox("👁 Vista Previa")
        layout = QVBoxLayout()
        
        self.label_resumen = QLabel("Cargue los archivos para ver el resumen")
        self.label_resumen.setStyleSheet("padding: 10px; color: #666;")
        layout.addWidget(self.label_resumen)
        
        grupo.setLayout(layout)
        return grupo
    
    def _cargar_excel(self):
        """Carga el archivo Excel"""
        archivo, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar archivo Excel",
            self.config.get_last_file_path('excel'),
            "Archivos Excel (*.xlsx *.xls *.xlsm)"
        )
        
        if not archivo:
            return
        
        self.excel_processor = ExcelProcessor(self.logger)
        exito, mensaje, clientes = self.excel_processor.procesar_archivo(archivo)
        
        if exito:
            self.clientes = clientes
            self.label_excel.setText(f"✅ {os.path.basename(archivo)} ({len(clientes)} clientes)")
            self.label_excel.setStyleSheet("color: green;")
            self.config.remember_file_path('excel', archivo)
            self._actualizar_preview()
        else:
            QMessageBox.critical(self, "Error al cargar Excel", mensaje)
            self.label_excel.setText("❌ Error al cargar Excel")
            self.label_excel.setStyleSheet("color: red;")
    
    def _cargar_zip(self):
        """Carga el archivo ZIP"""
        archivo, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar archivo ZIP",
            self.config.get_last_file_path('zip'),
            "Archivos ZIP (*.zip)"
        )
        
        if not archivo:
            return
        
        self.zip_handler = ZipHandler(self.logger)
        exito, mensaje, archivos_por_nit = self.zip_handler.procesar_zip(archivo)
        
        if exito:
            self.archivos_por_nit = archivos_por_nit
            total_archivos = sum(len(archivos) for archivos in archivos_por_nit.values())
            self.label_zip.setText(f"✅ {os.path.basename(archivo)} ({total_archivos} PDFs)")
            self.label_zip.setStyleSheet("color: green;")
            self.config.remember_file_path('zip', archivo)
            self._actualizar_preview()
        else:
            QMessageBox.critical(self, "Error al cargar ZIP", mensaje)
            self.label_zip.setText("❌ Error al cargar ZIP")
            self.label_zip.setStyleSheet("color: red;")
    
    def _actualizar_preview(self):
        """Actualiza la vista previa"""
        if not self.clientes or not self.archivos_por_nit:
            return
        
        # Validar coincidencias
        validacion = self.zip_handler.validar_archivos_contra_clientes(self.clientes)
        
        # Construir mensaje HTML con enlaces
        texto = f"""
<b>Resumen:</b><br>
• Clientes en Excel: {validacion['total_excel']}<br>
• PDFs en ZIP: {validacion['total_zip']}<br>
• Coincidencias: {validacion['coincidentes']}<br>
"""
        
        if validacion['sin_archivos']:
            texto += f"<br><b style='color: orange;'>⚠ {len(validacion['sin_archivos'])} clientes sin archivos</b>"
            texto += " <a href='#sin_archivos' style='color: blue;'>[Ver detalles]</a>"
        
        if validacion['sin_cliente']:
            texto += f"<br><b style='color: orange;'>⚠ {len(validacion['sin_cliente'])} archivos sin cliente</b>"
            texto += " <a href='#sin_cliente' style='color: blue;'>[Ver detalles]</a>"
        
        self.label_resumen.setText(texto)
        self.label_resumen.linkActivated.connect(lambda link: self._mostrar_detalles(link, validacion))
        
        # Habilitar botón si hay coincidencias
        if validacion['coincidentes'] > 0:
            self.btn_enviar.setEnabled(True)
        else:
            self.btn_enviar.setEnabled(False)
    
    def _mostrar_detalles(self, link, validacion):
        """Muestra diálogo con detalles de errores"""
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QTextEdit, QPushButton, QHBoxLayout
        
        dialogo = QDialog(self)
        dialogo.setWindowTitle("Detalles de Validación")
        dialogo.setMinimumSize(900, 600)
        
        layout = QVBoxLayout(dialogo)
        
        # Área de texto
        texto = QTextEdit()
        texto.setReadOnly(True)
        
        datos_exportar = None
        tipo_datos = ""
        
        if link == "#sin_archivos":
            tipo_datos = "clientes_sin_archivos"
            contenido = "<h2>📋 Clientes sin Archivos</h2>"
            contenido += f"<p>Total: {len(validacion['sin_archivos'])} clientes</p>"
            contenido += "<table border='1' cellpadding='5' style='border-collapse: collapse; width: 100%;'>"
            contenido += "<tr style='background-color: #f0f0f0;'><th>#</th><th>NIT</th><th>Nombre</th><th>Email</th></tr>"
            
            for i, cliente in enumerate(validacion['sin_archivos'], 1):
                contenido += f"<tr>"
                contenido += f"<td>{i}</td>"
                contenido += f"<td>{cliente['nit']}</td>"
                contenido += f"<td>{cliente['nombre']}</td>"
                contenido += f"<td>{cliente['email']}</td>"
                contenido += "</tr>"
            
            contenido += "</table>"
            texto.setHtml(contenido)
            datos_exportar = validacion['sin_archivos']
        
        elif link == "#sin_cliente":
            tipo_datos = "archivos_sin_cliente"
            contenido = "<h2>📁 Archivos sin Cliente</h2>"
            contenido += f"<p>Total: {len(validacion['sin_cliente'])} archivos</p>"
            contenido += "<table border='1' cellpadding='5' style='border-collapse: collapse; width: 100%;'>"
            contenido += "<tr style='background-color: #f0f0f0;'><th>#</th><th>NIT</th><th>Archivo</th></tr>"
            
            for i, info in enumerate(validacion['sin_cliente'], 1):
                contenido += f"<tr>"
                contenido += f"<td>{i}</td>"
                contenido += f"<td>{info['nit']}</td>"
                contenido += f"<td style='font-size: 10px;'>{info['archivo']}</td>"
                contenido += "</tr>"
            
            contenido += "</table>"
            texto.setHtml(contenido)
            datos_exportar = validacion['sin_cliente']
        
        layout.addWidget(texto)
        
        # Botones
        btn_layout = QHBoxLayout()
        
        btn_exportar = QPushButton("📊 Exportar a Excel")
        btn_exportar.clicked.connect(lambda: self._exportar_validacion(datos_exportar, tipo_datos))
        btn_layout.addWidget(btn_exportar)
        
        btn_layout.addStretch()
        
        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.clicked.connect(dialogo.close)
        btn_layout.addWidget(btn_cerrar)
        
        layout.addLayout(btn_layout)
        
        dialogo.exec()
    
    def _exportar_validacion(self, datos, tipo):
        """Exporta los datos de validación a Excel"""
        if not datos:
            return
        
        from datetime import datetime
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if tipo == "clientes_sin_archivos":
            ws.title = "Clientes sin Archivos"
            ws.append(['#', 'NIT', 'Nombre', 'Email'])
            
            # Estilo de encabezados
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            for i, cliente in enumerate(datos, 1):
                ws.append([i, cliente['nit'], cliente['nombre'], cliente['email']])
        
        elif tipo == "archivos_sin_cliente":
            ws.title = "Archivos sin Cliente"
            ws.append(['#', 'NIT', 'Archivo'])
            
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
            
            for i, info in enumerate(datos, 1):
                ws.append([i, info['nit'], info['archivo']])
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
        if tipo == "clientes_sin_archivos":
            ws.column_dimensions['D'].width = 40
        
        # Guardar
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"validacion_{tipo}_{timestamp}.xlsx"
        
        archivo, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar reporte",
            nombre_archivo,
            "Archivos Excel (*.xlsx)"
        )
        
        if archivo:
            try:
                wb.save(archivo)
                QMessageBox.information(self, "Exportado", f"Archivo guardado:\n{archivo}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al guardar: {str(e)}")
    
    def _enviar_comprobantes(self):
        """Inicia el envío de comprobantes"""
        from app.utils.validator import Validator
        
        # Validar credenciales SMTP
        correo = self.txt_correo.text().strip()
        password = self.txt_password_envio.text()
        servidor = self.txt_servidor.text().strip()
        puerto = self.spin_puerto.value()
        
        es_valido, mensaje = Validator.validar_configuracion_smtp(servidor, puerto, correo, password)
        if not es_valido:
            QMessageBox.warning(self, "Configuración inválida", 
                              f"Complete correctamente los datos de correo:\n{mensaje}")
            return
        
        # Preparar configuración SMTP
        smtp_config = {
            'server': servidor,
            'port': puerto,
            'username': correo,
            'password': password,
            'use_tls': True,
            'from_name': 'Sistema de Comprobantes'
        }
        
        # Obtener emails en copia
        texto_cc = self.txt_cc.text().strip()
        emails_cc = []
        if texto_cc:
            es_valido, mensaje, emails_cc = Validator.validar_lista_emails(texto_cc)
            if not es_valido:
                respuesta = QMessageBox.question(
                    self, "Emails CC inválidos",
                    f"{mensaje}\n\n¿Continuar sin los emails inválidos?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if respuesta == QMessageBox.StandardButton.No:
                    return
        
        # Preparar envíos (búsqueda flexible)
        clientes_con_archivos = []
        clientes_sin_archivos = []
        
        for cliente in self.clientes:
            nit = cliente['nit']
            nombre = cliente['nombre']
            
            # Primero buscar coincidencia exacta
            archivos = self.zip_handler.obtener_archivos_por_nit(nit)
            
            # Si no hay coincidencia exacta, buscar flexible (con nombre)
            if not archivos:
                archivos = self.zip_handler.buscar_archivos_por_nit_flexible(nit, nombre)
            
            if archivos:
                clientes_con_archivos.append({
                    'nit': nit,
                    'nombre': nombre,
                    'email': cliente['email'],
                    'archivos': archivos
                })
            else:
                clientes_sin_archivos.append(nombre)
        
        if not clientes_con_archivos:
            QMessageBox.warning(self, "Sin envíos", "No hay clientes con archivos para enviar")
            return
        
        # Mostrar advertencia si hay clientes sin archivos
        if clientes_sin_archivos:
            respuesta = QMessageBox.question(
                self, "Clientes sin archivos",
                f"⚠ {len(clientes_sin_archivos)} cliente(s) no tienen archivos:\n" +
                "\n".join(clientes_sin_archivos[:5]) +
                ("\n..." if len(clientes_sin_archivos) > 5 else "") +
                f"\n\n¿Continuar con los {len(clientes_con_archivos)} que sí tienen archivos?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if respuesta != QMessageBox.StandardButton.Yes:
                return
        
        # Confirmar
        modo_prueba = self.config.is_test_mode()
        mensaje_modo = " (MODO PRUEBA)" if modo_prueba else ""
        
        respuesta = QMessageBox.question(
            self, "Confirmar envío",
            f"¿Enviar comprobantes a {len(clientes_con_archivos)} clientes?{mensaje_modo}\n\n"
            f"Desde: {correo}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if respuesta != QMessageBox.StandardButton.Yes:
            return
        
        # Iniciar envío en thread
        self.btn_enviar.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setMaximum(len(clientes_con_archivos))
        self.progress_bar.setValue(0)
        self.txt_log.clear()
        
        self.worker = EnvioWorker(
            clientes_con_archivos,
            smtp_config,
            emails_cc,
            modo_prueba,
            self.db,
            self.logger
        )
        
        self.worker.progreso.connect(self._on_progreso)
        self.worker.finalizado.connect(self._on_finalizado)
        self.worker.start()
    
    def _on_progreso(self, actual, total, cliente, exito, estado, mensaje):
        """Callback de progreso"""
        self.progress_bar.setValue(actual)
        
        icono = "✅" if exito else "❌"
        self.txt_log.append(f"{icono} [{actual}/{total}] {cliente['nombre']} ({cliente['nit']}) - {estado}")
    
    def _on_finalizado(self, resumen):
        """Callback de finalización"""
        self.progress_bar.setVisible(False)
        self.btn_enviar.setEnabled(True)
        
        mensaje = f"""
<b>Envío completado</b><br>
<br>
✅ Enviados: {resumen['enviados']}<br>
❌ Errores: {resumen['errores']}<br>
🔄 Rebotados: {resumen['rebotados']}<br>
🚫 Bloqueados: {resumen['bloqueados']}<br>
❓ Inexistentes: {resumen['inexistentes']}<br>
"""
        
        QMessageBox.information(self, "Envío completado", mensaje)
        
        # Limpiar archivos temporales
        if self.zip_handler:
            self.zip_handler.limpiar_temporales()